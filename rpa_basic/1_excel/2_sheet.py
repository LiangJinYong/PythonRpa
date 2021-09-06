from openpyxl import Workbook

wb = Workbook()
# wb.active
ws = wb.create_sheet() # 새로운 Sheet 기본 이름으로 생성
ws.title = "PC Agent" # Sheet 이름 변경
ws.sheet_properties.tabColor = "ff66cc" # RGB 형태로 값을 넣어주면 탭 색상 변경

# Sheet, PC Agent, Device
ws1 = wb.create_sheet("Device") # 주어진 이름으로 Sheet 생성

ws2 = wb.create_sheet("Collector", 1) # 1번째 index에 Sheet 생성

device_ws = wb["Device"] # Dict 형태로 Sheet 에 접근

print(wb.sheetnames) # 모든 Sheet 이름 확인

# Sheet 복사
device_ws["A1"] = "Orange PI"
target = wb.copy_worksheet(device_ws)
target.title = "Copied Device"

wb.save("sample.xlsx")
wb.close()
